"""Parties, rollup reports and exports.

The rollups are what the broker actually opens each morning: who owes what,
which bills are unread, how much brokerage has accrued this year.
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import date

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.orm import Session, selectinload

from app.api import serializers as ser
from app.db import get_db
from app.extraction.normalize import clean_gstin, clean_text, normalize_name
from app.models import (
    BrokerageEntry,
    Document,
    Invoice,
    InvoiceLine,
    Party,
    PartyAlias,
    Product,
    ValidationFlag,
)

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api")

# Bills still being read, or already rejected, must not move the money totals.
COUNTED = ("extracted", "needs_review", "confirmed")


def _counted():
    return Invoice.status.in_(COUNTED)


def _date_filters(stmt, date_from: date | None, date_to: date | None, fy: str | None):
    if date_from:
        stmt = stmt.where(Invoice.invoice_date >= date_from)
    if date_to:
        stmt = stmt.where(Invoice.invoice_date <= date_to)
    if fy:
        stmt = stmt.where(Invoice.financial_year == fy)
    return stmt


# ==========================================================================
# Parties
# ==========================================================================


@router.get("/parties")
def list_parties(
    role: str | None = Query(None, pattern="^(seller|buyer|transporter|broker)$"),
    q: str | None = None,
    limit: int = Query(200, le=1000),
    offset: int = 0,
    db: Session = Depends(get_db),
) -> dict:
    stmt = select(Party).options(selectinload(Party.aliases))
    if role:
        stmt = stmt.where(getattr(Party, f"is_{role}").is_(True))
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            Party.legal_name.ilike(like)
            | Party.gstin.ilike(like)
            | Party.city.ilike(like)
        )
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(
        stmt.order_by(Party.legal_name).limit(limit).offset(offset)
    ).unique().all()
    return {"total": total, "parties": [ser.party_full(p) for p in rows]}


@router.get("/parties/{party_id}")
def get_party(party_id: int, db: Session = Depends(get_db)) -> dict:
    party = db.get(Party, party_id)
    if party is None:
        raise HTTPException(404, "Party not found")

    data = ser.party_full(party)
    for role, column in (
        ("as_seller", Invoice.seller_id),
        ("as_buyer", Invoice.buyer_id),
        ("as_transporter", Invoice.transporter_id),
        ("as_broker", Invoice.broker_id),
    ):
        row = db.execute(
            select(
                func.count(Invoice.id),
                func.coalesce(func.sum(Invoice.grand_total), 0),
                func.coalesce(func.sum(Invoice.taxable_value), 0),
                func.min(Invoice.invoice_date),
                func.max(Invoice.invoice_date),
            ).where(column == party_id, _counted())
        ).one()
        data[role] = {
            "invoice_count": row[0],
            "grand_total": float(row[1] or 0),
            "taxable_value": float(row[2] or 0),
            "first_invoice": ser.iso(row[3]),
            "last_invoice": ser.iso(row[4]),
        }
    return data


@router.patch("/parties/{party_id}")
def update_party(party_id: int, payload: dict = Body(...), db: Session = Depends(get_db)) -> dict:
    party = db.get(Party, party_id)
    if party is None:
        raise HTTPException(404, "Party not found")

    editable = {
        "legal_name": clean_text, "display_name": clean_text, "gstin": clean_gstin,
        "pan": clean_text, "fssai": clean_text, "address": clean_text,
        "city": clean_text, "state_name": clean_text, "state_code": clean_text,
        "pincode": clean_text, "phone": clean_text, "email": clean_text,
        "contact_person": clean_text, "notes": clean_text,
    }
    for field, cleaner in editable.items():
        if field in payload:
            setattr(party, field, cleaner(payload[field]))
    for flag in ("is_seller", "is_buyer", "is_transporter", "is_broker", "is_verified"):
        if flag in payload:
            setattr(party, flag, bool(payload[flag]))
    for field in ("credit_days",):
        if field in payload:
            setattr(party, field, int(payload[field]) if payload[field] else None)

    if payload.get("legal_name"):
        party.normalized_name = normalize_name(party.legal_name)
    db.commit()
    return ser.party_full(party)


@router.post("/parties/merge")
def merge_parties(payload: dict = Body(...), db: Session = Depends(get_db)) -> dict:
    """Fold a duplicate party into the one to keep.

    Matching errs towards creating a duplicate rather than fusing two real
    firms, so this is the intended cleanup path — and it is deliberately
    manual, because only a human knows whether two similar names are one
    business.
    """
    keep_id, drop_id = payload.get("keep_id"), payload.get("drop_id")
    if not keep_id or not drop_id or keep_id == drop_id:
        raise HTTPException(400, "Provide two different party ids: keep_id and drop_id.")

    keep, drop = db.get(Party, keep_id), db.get(Party, drop_id)
    if keep is None or drop is None:
        raise HTTPException(404, "One of the parties does not exist.")
    if keep.gstin and drop.gstin and keep.gstin != drop.gstin:
        raise HTTPException(
            409,
            f"These have different GSTINs ({keep.gstin} and {drop.gstin}), so they "
            "are different legal entities. Merge refused.",
        )

    moved = 0
    for column in (Invoice.seller_id, Invoice.buyer_id, Invoice.consignee_id,
                   Invoice.transporter_id, Invoice.broker_id):
        for inv in db.scalars(select(Invoice).where(column == drop_id)).all():
            setattr(inv, column.key, keep_id)
            moved += 1

    for role in ("is_seller", "is_buyer", "is_transporter", "is_broker"):
        if getattr(drop, role):
            setattr(keep, role, True)
    if drop.gstin and not keep.gstin:
        keep.gstin = drop.gstin

    # Keep the dropped name as an alias so the next bill spelled that way
    # lands on the surviving party.
    existing = {a.normalized_alias for a in keep.aliases}
    for alias_text in [drop.legal_name] + [a.alias for a in drop.aliases]:
        key = normalize_name(alias_text)
        if key and key not in existing and key != keep.normalized_name:
            db.add(PartyAlias(party_id=keep.id, alias=alias_text,
                              normalized_alias=key, source="merge"))
            existing.add(key)

    db.delete(drop)
    db.commit()
    return {"kept": keep_id, "dropped": drop_id, "invoices_moved": moved}


# ==========================================================================
# Rollups
# ==========================================================================


def _rollup(db: Session, column, *, date_from, date_to, fy, limit: int) -> list[dict]:
    stmt = (
        select(
            column.label("party_id"),
            func.count(Invoice.id).label("invoice_count"),
            func.coalesce(func.sum(Invoice.taxable_value), 0).label("taxable_value"),
            func.coalesce(func.sum(Invoice.grand_total), 0).label("grand_total"),
            func.coalesce(func.sum(Invoice.amount_paid), 0).label("amount_paid"),
            func.max(Invoice.invoice_date).label("last_invoice"),
        )
        .where(column.isnot(None), _counted())
        .group_by(column)
        .order_by(func.sum(Invoice.grand_total).desc().nullslast())
        .limit(limit)
    )
    stmt = _date_filters(stmt, date_from, date_to, fy)

    rows = db.execute(stmt).all()
    parties = {
        p.id: p
        for p in db.scalars(select(Party).where(Party.id.in_([r.party_id for r in rows]))).all()
    }
    out = []
    for row in rows:
        party = parties.get(row.party_id)
        grand = float(row.grand_total or 0)
        paid = float(row.amount_paid or 0)
        out.append({
            "party": ser.party_brief(party),
            "invoice_count": row.invoice_count,
            "taxable_value": float(row.taxable_value or 0),
            "grand_total": grand,
            "amount_paid": paid,
            "outstanding": round(grand - paid, 2),
            "last_invoice": ser.iso(row.last_invoice),
        })
    return out


# Declared before the generic by-{dimension} route below: FastAPI matches
# in registration order, and the path parameter would otherwise swallow
# "product" and reject it as an unknown grouping.
@router.get("/reports/by-product")
def product_report(
    date_from: date | None = None,
    date_to: date | None = None,
    financial_year: str | None = None,
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
) -> dict:
    """Commodity-wise volume and value, with a weighted average rate."""
    stmt = (
        select(
            Product.id,
            Product.canonical_name,
            Product.category,
            Product.default_hsn,
            func.count(func.distinct(Invoice.id)).label("invoice_count"),
            func.coalesce(func.sum(InvoiceLine.quantity), 0).label("quantity"),
            func.coalesce(func.sum(InvoiceLine.taxable_amount), 0).label("value"),
        )
        .select_from(InvoiceLine)
        .join(Invoice, Invoice.id == InvoiceLine.invoice_id)
        .join(Product, Product.id == InvoiceLine.product_id)
        .where(_counted())
        .group_by(Product.id, Product.canonical_name, Product.category, Product.default_hsn)
        .order_by(func.sum(InvoiceLine.taxable_amount).desc().nullslast())
        .limit(limit)
    )
    stmt = _date_filters(stmt, date_from, date_to, financial_year)

    rows = []
    for r in db.execute(stmt).all():
        qty = float(r.quantity or 0)
        value = float(r.value or 0)
        rows.append({
            "product_id": r.id,
            "product": r.canonical_name,
            "category": r.category,
            "hsn": r.default_hsn,
            "invoice_count": r.invoice_count,
            "quantity": qty,
            "value": value,
            "avg_rate": round(value / qty, 2) if qty else None,
        })
    return {"dimension": "product", "rows": rows}


@router.get("/reports/by-{dimension}")
def rollup_report(
    dimension: str,
    date_from: date | None = None,
    date_to: date | None = None,
    financial_year: str | None = None,
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
) -> dict:
    """Turnover grouped by seller, buyer, transporter or broker."""
    columns = {
        "seller": Invoice.seller_id,
        "buyer": Invoice.buyer_id,
        "transporter": Invoice.transporter_id,
        "broker": Invoice.broker_id,
    }
    if dimension not in columns:
        raise HTTPException(
            404, f"Unknown grouping '{dimension}'. Use one of: {', '.join(columns)}."
        )
    return {
        "dimension": dimension,
        "rows": _rollup(
            db, columns[dimension],
            date_from=date_from, date_to=date_to, fy=financial_year, limit=limit,
        ),
    }


@router.get("/reports/summary")
def summary(
    financial_year: str | None = None,
    db: Session = Depends(get_db),
) -> dict:
    """Dashboard headline numbers."""
    stmt = select(
        func.count(Invoice.id),
        func.coalesce(func.sum(Invoice.grand_total), 0),
        func.coalesce(func.sum(Invoice.taxable_value), 0),
        # Each column is coalesced separately: in SQL NULL + 5 is NULL, so
        # summing the raw expression yields nothing whenever a bill carries
        # only CGST/SGST or only IGST — which is every bill.
        func.coalesce(func.sum(func.coalesce(Invoice.cgst_amount, 0)), 0)
        + func.coalesce(func.sum(func.coalesce(Invoice.sgst_amount, 0)), 0)
        + func.coalesce(func.sum(func.coalesce(Invoice.igst_amount, 0)), 0),
        func.coalesce(func.sum(Invoice.amount_paid), 0),
    ).where(_counted())
    if financial_year:
        stmt = stmt.where(Invoice.financial_year == financial_year)
    count, grand, taxable, tax, paid = db.execute(stmt).one()

    doc_counts = dict(
        db.execute(select(Document.status, func.count(Document.id)).group_by(Document.status)).all()
    )
    inv_counts = dict(
        db.execute(select(Invoice.status, func.count(Invoice.id)).group_by(Invoice.status)).all()
    )

    brokerage_stmt = select(func.coalesce(func.sum(BrokerageEntry.amount), 0))
    if financial_year:
        brokerage_stmt = brokerage_stmt.where(BrokerageEntry.financial_year == financial_year)
    brokerage = db.scalar(brokerage_stmt) or 0

    years = [
        r[0] for r in db.execute(
            select(Invoice.financial_year)
            .where(Invoice.financial_year.isnot(None))
            .group_by(Invoice.financial_year)
            .order_by(Invoice.financial_year.desc())
        ).all()
    ]

    return {
        "financial_year": financial_year,
        "available_years": years,
        "invoice_count": count or 0,
        "grand_total": float(grand or 0),
        "taxable_value": float(taxable or 0),
        "tax_total": float(tax or 0),
        "amount_paid": float(paid or 0),
        "outstanding": float((grand or 0) - (paid or 0)),
        "brokerage_accrued": float(brokerage),
        "documents_by_status": doc_counts,
        "invoices_by_status": inv_counts,
        "needs_review": inv_counts.get("needs_review", 0),
        "party_count": db.scalar(select(func.count(Party.id))) or 0,
        "product_count": db.scalar(select(func.count(Product.id))) or 0,
    }


@router.get("/reports/timeline")
def timeline(
    financial_year: str | None = None,
    db: Session = Depends(get_db),
) -> dict:
    """Turnover and brokerage month by month.

    Grouped in Python rather than in SQL: `date_trunc` is Postgres, `strftime`
    is SQLite, and this app is routinely run on both. The row count here is
    one per invoice, which is nothing to carry into memory.
    """
    stmt = select(
        Invoice.invoice_date, Invoice.grand_total, Invoice.taxable_value,
    ).where(_counted(), Invoice.invoice_date.is_not(None))
    if financial_year:
        stmt = stmt.where(Invoice.financial_year == financial_year)

    buckets: dict[str, dict] = {}
    for invoice_date, grand, taxable in db.execute(stmt).all():
        key = invoice_date.strftime("%Y-%m")
        row = buckets.setdefault(
            key, {"month": key, "label": invoice_date.strftime("%b %Y"),
                  "invoice_count": 0, "grand_total": 0.0, "taxable_value": 0.0}
        )
        row["invoice_count"] += 1
        row["grand_total"] += float(grand or 0)
        row["taxable_value"] += float(taxable or 0)

    brokerage = select(
        Invoice.invoice_date, func.coalesce(BrokerageEntry.amount, 0)
    ).join(BrokerageEntry, BrokerageEntry.invoice_id == Invoice.id).where(
        _counted(), Invoice.invoice_date.is_not(None)
    )
    if financial_year:
        brokerage = brokerage.where(Invoice.financial_year == financial_year)
    for invoice_date, amount in db.execute(brokerage).all():
        key = invoice_date.strftime("%Y-%m")
        if key in buckets:
            buckets[key]["brokerage"] = buckets[key].get("brokerage", 0.0) + float(amount or 0)

    months = [dict(b, brokerage=b.get("brokerage", 0.0))
              for b in sorted(buckets.values(), key=lambda r: r["month"])]
    return {"financial_year": financial_year, "months": months}


@router.get("/reports/review-queue")
def review_queue(limit: int = Query(100, le=500), db: Session = Depends(get_db)) -> dict:
    """Invoices waiting on a human, worst first."""
    stmt = (
        select(Invoice)
        .options(
            selectinload(Invoice.seller),
            selectinload(Invoice.buyer),
            selectinload(Invoice.transporter),
            selectinload(Invoice.broker),
            selectinload(Invoice.flags),
        )
        .where(Invoice.needs_review.is_(True))
        .order_by(Invoice.confidence.asc().nullsfirst(), Invoice.created_at.desc())
        .limit(limit)
    )
    rows = db.scalars(stmt).unique().all()
    out = []
    for inv in rows:
        data = ser.invoice_brief(inv)
        data["flags"] = [ser.flag(f) for f in inv.flags if not f.resolved]
        out.append(data)

    failed = db.scalars(
        select(Document).where(Document.status == "failed").order_by(Document.created_at.desc())
    ).all()
    return {
        "invoices": out,
        "failed_documents": [ser.document(d) for d in failed],
    }


@router.get("/reports/corrections")
def corrections_report(limit: int = Query(50, le=500), db: Session = Depends(get_db)) -> dict:
    """Which fields humans keep fixing, by bill format.

    A field near the top of this list is an extraction problem to fix at the
    prompt, not a chore to keep repeating.
    """
    from app.models import Correction

    rows = db.execute(
        select(
            Correction.vendor_format_hint,
            Correction.field_path,
            func.count(Correction.id).label("n"),
        )
        .group_by(Correction.vendor_format_hint, Correction.field_path)
        .order_by(func.count(Correction.id).desc())
        .limit(limit)
    ).all()
    return {
        "rows": [
            {"vendor_format": r[0] or "unknown", "field": r[1], "corrections": r[2]}
            for r in rows
        ]
    }


# ==========================================================================
# Exports
# ==========================================================================

INVOICE_COLUMNS = [
    ("invoice_id", lambda i: i.id),
    ("invoice_number", lambda i: i.invoice_number),
    ("invoice_date", lambda i: ser.iso(i.invoice_date)),
    ("financial_year", lambda i: i.financial_year),
    ("document_type", lambda i: i.document_type),
    ("seller", lambda i: i.seller.legal_name if i.seller else None),
    ("seller_gstin", lambda i: i.seller.gstin if i.seller else None),
    ("buyer", lambda i: i.buyer.legal_name if i.buyer else None),
    ("buyer_gstin", lambda i: i.buyer.gstin if i.buyer else None),
    ("consignee", lambda i: i.consignee.legal_name if i.consignee else None),
    ("broker", lambda i: (i.broker.legal_name if i.broker else i.broker_name_raw)),
    ("transporter", lambda i: i.transporter.legal_name if i.transporter else None),
    ("vehicle_no", lambda i: i.vehicle_no),
    ("eway_bill_no", lambda i: i.eway_bill_no),
    ("place_of_supply", lambda i: i.place_of_supply),
    ("supply_type", lambda i: i.supply_type),
    ("payment_terms", lambda i: i.payment_terms),
    ("due_date", lambda i: ser.iso(i.due_date)),
    ("total_quantity", lambda i: ser.num(i.total_quantity)),
    ("uom", lambda i: i.total_quantity_uom),
    ("total_bags", lambda i: ser.num(i.total_bags)),
    ("subtotal", lambda i: ser.num(i.subtotal)),
    ("discount_total", lambda i: ser.num(i.discount_total)),
    ("taxable_value", lambda i: ser.num(i.taxable_value)),
    ("cgst", lambda i: ser.num(i.cgst_amount)),
    ("sgst", lambda i: ser.num(i.sgst_amount)),
    ("igst", lambda i: ser.num(i.igst_amount)),
    ("cess", lambda i: ser.num(i.cess_amount)),
    ("tcs", lambda i: ser.num(i.tcs_amount)),
    ("other_charges", lambda i: ser.num(i.other_charges)),
    ("round_off", lambda i: ser.num(i.round_off)),
    ("grand_total", lambda i: ser.num(i.grand_total)),
    ("amount_paid", lambda i: ser.num(i.amount_paid)),
    ("payment_status", lambda i: i.payment_status),
    ("status", lambda i: i.status),
    ("confidence", lambda i: i.confidence),
    ("irn", lambda i: i.irn),
    ("source_file", lambda i: i.document.original_filename if i.document else None),
]

LINE_COLUMNS = [
    ("invoice_id", lambda i, l: i.id),
    ("invoice_number", lambda i, l: i.invoice_number),
    ("invoice_date", lambda i, l: ser.iso(i.invoice_date)),
    ("seller", lambda i, l: i.seller.legal_name if i.seller else None),
    ("buyer", lambda i, l: i.buyer.legal_name if i.buyer else None),
    ("line_no", lambda i, l: l.line_no),
    ("description", lambda i, l: l.description),
    ("item_code", lambda i, l: l.item_code),
    ("grade", lambda i, l: l.item_remarks),
    ("brand", lambda i, l: l.brand),
    ("product", lambda i, l: l.product.canonical_name if l.product else None),
    ("hsn", lambda i, l: l.hsn),
    ("bags", lambda i, l: ser.num(l.bags)),
    ("quantity", lambda i, l: ser.num(l.quantity)),
    ("uom", lambda i, l: l.uom),
    ("rate", lambda i, l: ser.num(l.rate)),
    ("discount_pct", lambda i, l: ser.num(l.discount_pct)),
    ("taxable_amount", lambda i, l: ser.num(l.taxable_amount)),
    ("tax_rate", lambda i, l: ser.num(l.tax_rate)),
    ("cgst", lambda i, l: ser.num(l.cgst_amount)),
    ("sgst", lambda i, l: ser.num(l.sgst_amount)),
    ("igst", lambda i, l: ser.num(l.igst_amount)),
]


def _export_query(db: Session, date_from, date_to, fy, seller_id, buyer_id, status):
    stmt = (
        select(Invoice)
        .options(
            selectinload(Invoice.seller), selectinload(Invoice.buyer),
            selectinload(Invoice.consignee), selectinload(Invoice.transporter),
            selectinload(Invoice.broker), selectinload(Invoice.document),
            selectinload(Invoice.lines).selectinload(InvoiceLine.product),
        )
        .order_by(Invoice.invoice_date, Invoice.id)
    )
    stmt = _date_filters(stmt, date_from, date_to, fy)
    if seller_id:
        stmt = stmt.where(Invoice.seller_id == seller_id)
    if buyer_id:
        stmt = stmt.where(Invoice.buyer_id == buyer_id)
    stmt = stmt.where(Invoice.status == status) if status else stmt.where(_counted())
    return db.scalars(stmt).unique().all()


@router.get("/exports/invoices.csv")
def export_invoices_csv(
    date_from: date | None = None,
    date_to: date | None = None,
    financial_year: str | None = None,
    seller_id: int | None = None,
    buyer_id: int | None = None,
    status: str | None = None,
    level: str = Query("invoice", pattern="^(invoice|line)$"),
    db: Session = Depends(get_db),
):
    invoices = _export_query(db, date_from, date_to, financial_year, seller_id, buyer_id, status)
    buf = io.StringIO()
    writer = csv.writer(buf)

    if level == "invoice":
        writer.writerow([c[0] for c in INVOICE_COLUMNS])
        for inv in invoices:
            writer.writerow([fn(inv) for _, fn in INVOICE_COLUMNS])
        name = "invoices.csv"
    else:
        writer.writerow([c[0] for c in LINE_COLUMNS])
        for inv in invoices:
            for row in inv.lines:
                writer.writerow([fn(inv, row) for _, fn in LINE_COLUMNS])
        name = "invoice_lines.csv"

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.get("/exports/invoices.xlsx")
def export_invoices_xlsx(
    date_from: date | None = None,
    date_to: date | None = None,
    financial_year: str | None = None,
    seller_id: int | None = None,
    buyer_id: int | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
):
    """Workbook with an Invoices sheet and a Line Items sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    invoices = _export_query(db, date_from, date_to, financial_year, seller_id, buyer_id, status)
    wb = Workbook()

    sheet = wb.active
    sheet.title = "Invoices"
    sheet.append([c[0] for c in INVOICE_COLUMNS])
    for inv in invoices:
        sheet.append([fn(inv) for _, fn in INVOICE_COLUMNS])

    lines_sheet = wb.create_sheet("Line Items")
    lines_sheet.append([c[0] for c in LINE_COLUMNS])
    for inv in invoices:
        for row in inv.lines:
            lines_sheet.append([fn(inv, row) for _, fn in LINE_COLUMNS])

    for ws, columns in ((sheet, INVOICE_COLUMNS), (lines_sheet, LINE_COLUMNS)):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for idx, (name, _) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(len(name) + 6, 40))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="invoices.xlsx"'},
    )
